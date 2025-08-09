import streamlit as st
import pandas as pd
import numpy as np
import os
import json
from datetime import datetime
from tabulate import tabulate
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill, Font, Alignment, numbers
from openpyxl.utils import get_column_letter
from tqdm import tqdm
import io
import tempfile
import base64

# Configuração da página
st.set_page_config(
    page_title="🏆 Handicap/ML Pro",
    page_icon="🏆",
    layout="wide",
    initial_sidebar_state="expanded"
)

# CSS personalizado para deixar a interface bonita
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
    
    html, body, [class*="css"] {
        font-family: 'Inter', sans-serif;
    }
    
    .main-header {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 2rem;
        border-radius: 15px;
        color: white;
        text-align: center;
        margin-bottom: 2rem;
        box-shadow: 0 8px 32px rgba(102, 126, 234, 0.3);
    }
    
    .main-header h1 {
        font-size: 2.5rem;
        font-weight: 700;
        margin: 0;
        text-shadow: 2px 2px 4px rgba(0,0,0,0.3);
    }
    
    .main-header p {
        font-size: 1.2rem;
        margin: 0.5rem 0 0 0;
        opacity: 0.9;
    }
    
    .metric-card {
        background: linear-gradient(145deg, #f8fafc, #e2e8f0);
        padding: 1.5rem;
        border-radius: 12px;
        border-left: 4px solid #3b82f6;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.05);
        margin: 0.5rem 0;
    }
    
    .metric-value {
        font-size: 2rem;
        font-weight: 700;
        color: #1e40af;
        margin-bottom: 0.25rem;
    }
    
    .metric-label {
        font-size: 0.875rem;
        color: #6b7280;
        font-weight: 500;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }
    
    .section-header {
        font-size: 1.5rem;
        font-weight: 600;
        color: #374151;
        margin: 2rem 0 1rem 0;
        padding-bottom: 0.5rem;
        border-bottom: 2px solid #e5e7eb;
    }
    
    .config-section {
        background: #f8fafc;
        padding: 1.5rem;
        border-radius: 12px;
        border: 1px solid #e5e7eb;
        margin: 1rem 0;
    }
    
    .config-title {
        font-weight: 600;
        color: #4b5563;
        margin-bottom: 1rem;
        font-size: 1.1rem;
    }
    
    .stButton > button {
        background: linear-gradient(145deg, #3b82f6, #2563eb);
        color: white;
        border: none;
        border-radius: 8px;
        padding: 0.75rem 1.5rem;
        font-weight: 600;
        font-size: 1rem;
        transition: all 0.3s ease;
        box-shadow: 0 4px 6px rgba(59, 130, 246, 0.3);
    }
    
    .stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 12px rgba(59, 130, 246, 0.4);
    }
    
    .success-button > button {
        background: linear-gradient(145deg, #10b981, #059669);
        box-shadow: 0 4px 6px rgba(16, 185, 129, 0.3);
    }
    
    .danger-button > button {
        background: linear-gradient(145deg, #ef4444, #dc2626);
        box-shadow: 0 4px 6px rgba(239, 68, 68, 0.3);
    }
    
    .etapa-card {
        background: white;
        border: 1px solid #e5e7eb;
        border-radius: 8px;
        padding: 1rem;
        margin: 0.5rem 0;
        box-shadow: 0 2px 4px rgba(0, 0, 0, 0.05);
    }
    
    .roi-positive {
        color: #059669;
        font-weight: 600;
    }
    
    .roi-negative {
        color: #dc2626;
        font-weight: 600;
    }
    
    .status-indicator {
        padding: 0.25rem 0.75rem;
        border-radius: 20px;
        font-size: 0.875rem;
        font-weight: 500;
    }
    
    .status-processing {
        background: #dbeafe;
        color: #1e40af;
    }
    
    .status-success {
        background: #d1fae5;
        color: #065f46;
    }
    
    .status-error {
        background: #fecaca;
        color: #991b1b;
    }
</style>
""", unsafe_allow_html=True)

class BacktestAnalyzer:
    """Classe principal para análise de backtest com busca otimizada"""
    
    def __init__(self):
        self.reset_state()
        self.config_file = "busca_config_streamlit.json"
        
        # Configurações padrão da busca gulosa
        self.default_busca_config = {
            'usar_winrate1': True,
            'usar_winrate2': True,
            'usar_excl_campeonatos': True,
            'usar_excl_apostas_a_favor': True,
            'usar_excl_apostas_contra': True,
            'usar_excl_confrontos': True,
            'usar_excl_times_a_favor': True,
            'usar_excl_times_contra': True,
            'usar_excl_tipo_apostas': True,
            'usar_excl_tipo_local': True,
            'usar_diferenca_placar_min': True,
            'usar_diferenca_placar_max': True
        }
        
        # Configurações de quantidade mínima
        self.default_min_config = {
            'min_campeonatos': 10,
            'min_apostas_a_favor': 10,
            'min_apostas_contra': 10,
            'min_confrontos': 10,
            'min_times_a_favor': 10,
            'min_times_contra': 10,
            'min_tipo_apostas': 10,
            'min_tipo_local': 10,
            'min_winrate1': 10,
            'min_winrate2': 10,
            'min_diferenca_placar': 10
        }
        
        self.carregar_configuracoes()
    
    def reset_state(self):
        """Reset do estado da aplicação"""
        self.df = None
        self.df_filtrado = None
        self.torneio_escolhido = None
        self.campeonato_escolhido = None
        self.tip_escolhido = None
        self.roi_desejado = None
        self.etapas_filtros = []
        self.config = None
        self.melhor_df = None
        self.melhor_roi = None
        self.melhor_config = None
        self.total_inicial_apostas = 0
        self.limite_minimo_apostas = 0
        self.roi_inicial = 0
        
    def carregar_configuracoes(self):
        """Carregar configurações salvas"""
        try:
            if os.path.exists(self.config_file):
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    config_data = json.load(f)
                self.busca_config = config_data.get('busca_config', self.default_busca_config)
                self.min_entradas_config = config_data.get('min_entradas_config', self.default_min_config)
            else:
                self.busca_config = self.default_busca_config.copy()
                self.min_entradas_config = self.default_min_config.copy()
        except Exception as e:
            st.error(f"Erro ao carregar configurações: {e}")
            self.busca_config = self.default_busca_config.copy()
            self.min_entradas_config = self.default_min_config.copy()
    
    def salvar_configuracoes(self):
        """Salvar configurações atuais"""
        try:
            config_data = {
                'busca_config': self.busca_config,
                'min_entradas_config': self.min_entradas_config
            }
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(config_data, f, indent=2, ensure_ascii=False)
            return True
        except Exception as e:
            st.error(f"Erro ao salvar configurações: {e}")
            return False
    
    def carregar_arquivo(self, uploaded_file):
        """Carregar e processar arquivo Excel"""
        try:
            self.df = pd.read_excel(uploaded_file)
            
            # Tratar colunas numéricas
            if "Winrate 1" in self.df.columns:
                self.df["Winrate 1"] = pd.to_numeric(
                    self.df["Winrate 1"].astype(str).str.replace('%', ''), 
                    errors="coerce"
                )
            if "Winrate 2" in self.df.columns:
                self.df["Winrate 2"] = pd.to_numeric(
                    self.df["Winrate 2"].astype(str).str.replace('%', ''), 
                    errors="coerce"
                )
            if "Lucro/Prej." in self.df.columns:
                self.df["Lucro/Prej."] = pd.to_numeric(self.df["Lucro/Prej."], errors="coerce").fillna(0)
            
            # Verificar colunas necessárias
            required_columns = ["Torneio", "Jogador A", "Jogador B", "Tip", "Lucro/Prej.", "Winrate 1", "Winrate 2"]
            missing_columns = [col for col in required_columns if col not in self.df.columns]
            
            if missing_columns:
                st.error(f"❌ Colunas não encontradas: {', '.join(missing_columns)}")
                return False
            
            return True
            
        except Exception as e:
            st.error(f"❌ Erro ao carregar arquivo: {e}")
            return False
    
    def obter_opcoes_formulario(self):
        """Obter opções disponíveis para os formulários"""
        if self.df is None:
            return {}
        
        torneios = sorted(self.df['Torneio'].unique().tolist())
        torneios.append("Todos os torneios")
        
        return {'torneios': torneios}
    
    def obter_campeonatos(self, torneio_escolhido):
        """Obter campeonatos baseado no torneio escolhido"""
        if self.df is None:
            return ["Todos os campeonatos"]
        
        if torneio_escolhido == "Todos os torneios":
            df_torneio = self.df.copy()
        else:
            df_torneio = self.df[self.df['Torneio'] == torneio_escolhido].copy()
        
        campeonatos = sorted(df_torneio['Campeonato'].unique().tolist())
        campeonatos.append("Todos os campeonatos")
        
        return campeonatos
    
    def obter_tips_disponiveis(self, torneio_escolhido, campeonato_escolhido):
        """Obter tips disponíveis baseado nas seleções"""
        if self.df is None:
            return []
        
        df_temp = self.df.copy()
        
        if torneio_escolhido != "Todos os torneios":
            df_temp = df_temp[df_temp['Torneio'] == torneio_escolhido]
        
        if campeonato_escolhido != "Todos os campeonatos":
            df_temp = df_temp[df_temp['Campeonato'] == campeonato_escolhido]
        
        if 'Tip' in df_temp.columns:
            tips = df_temp['Tip'].unique().tolist()
            opcoes = []
            if 'Over' in tips:
                opcoes.append('Over')
            if 'Under' in tips:
                opcoes.append('Under')
            if len(opcoes) > 1:
                opcoes.append('Ambos')
            return opcoes
        
        return []
    
    def filtrar_dados_iniciais(self, torneio, campeonato, tip):
        """Filtrar dados baseado nas seleções iniciais"""
        if self.df is None:
            return False
        
        # Filtrar por torneio
        if torneio == "Todos os torneios":
            df_torneio = self.df.copy()
            self.torneio_escolhido = None
        else:
            df_torneio = self.df[self.df['Torneio'] == torneio].copy()
            self.torneio_escolhido = torneio
        
        # Filtrar por campeonato
        if campeonato == "Todos os campeonatos":
            self.df_filtrado = df_torneio.copy()
            self.campeonato_escolhido = None
        else:
            self.df_filtrado = df_torneio[df_torneio['Campeonato'] == campeonato].copy()
            self.campeonato_escolhido = campeonato
        
        # Filtrar por tip
        if tip and tip != "Ambos":
            if tip == 'Over':
                self.df_filtrado = self.df_filtrado[self.df_filtrado['Tip'] == 'Over'].copy()
                self.tip_escolhido = "Over"
            elif tip == 'Under':
                self.df_filtrado = self.df_filtrado[self.df_filtrado['Tip'] == 'Under'].copy()
                self.tip_escolhido = "Under"
        else:
            self.tip_escolhido = None
        
        # Adicionar coluna Confronto se não existir
        if 'Confronto' not in self.df_filtrado.columns:
            self.df_filtrado['Confronto'] = self.df_filtrado.apply(
                lambda row: f"{row['Jogador A']} vs {row['Jogador B']}" 
                if pd.notna(row["Jogador A"]) and pd.notna(row["Jogador B"]) else "", axis=1
            )
        
        return True
    
    def calcular_roi(self, df_atual):
        """Calcular ROI do dataframe atual"""
        if df_atual is None or len(df_atual) == 0:
            return -float('inf')
        
        lucro_total = df_atual['Lucro/Prej.'].sum()
        total_apostas = len(df_atual)
        return lucro_total / total_apostas if total_apostas > 0 else -float('inf')
    
    def calcular_diferenca_placar(self, placar):
        """Calcular diferença entre placares"""
        try:
            if pd.isna(placar) or placar is None or not isinstance(placar, str):
                return None
            if '-' not in placar:
                return None
            
            partes = placar.split('-')
            if len(partes) != 2:
                return None
            
            try:
                num1 = int(partes[0])
                num2 = int(partes[1])
                diferenca = abs(num1 - num2)
                return diferenca
            except (ValueError, TypeError):
                return None
        except Exception:
            return None
    
    def aplicar_filtros(self, df_base, config):
        """Aplicar filtros baseado na configuração"""
        df_temp = df_base.copy()
        
        # Extrair parâmetros da configuração
        w1 = config.get('w1')
        w2 = config.get('w2')
        apostas_a_favor_excl = config.get('apostas_a_favor_excl', [])
        apostas_contra_excl = config.get('apostas_contra_excl', [])
        confrontos_excl = config.get('confrontos', [])
        campeonatos_excl = config.get('campeonatos_excl', [])
        times_a_favor_excl = config.get('times_a_favor_excl', [])
        times_contra_excl = config.get('times_contra_excl', [])
        tipo_apostas_excl = config.get('tipo_apostas_excl', [])
        tipo_local_excl = config.get('tipo_local_excl', [])
        diferenca_placar_min = config.get('diferenca_placar_min')
        diferenca_placar_max = config.get('diferenca_placar_max')
        
        # Aplicar filtros
        if w1 is not None:
            df_temp = df_temp[df_temp['Winrate 1'] >= w1]
        if w2 is not None:
            df_temp = df_temp[df_temp['Winrate 2'] >= w2]
        if apostas_a_favor_excl:
            df_temp = df_temp[~((df_temp['Tip'] == df_temp['Jogador A']) & (df_temp['Jogador A'].isin(apostas_a_favor_excl))) &
                              ~((df_temp['Tip'] == df_temp['Jogador B']) & (df_temp['Jogador B'].isin(apostas_a_favor_excl)))]
        if apostas_contra_excl:
            df_temp = df_temp[~((df_temp['Tip'] != df_temp['Jogador A']) & (df_temp['Jogador A'].isin(apostas_contra_excl))) &
                              ~((df_temp['Tip'] != df_temp['Jogador B']) & (df_temp['Jogador B'].isin(apostas_contra_excl)))]
        if confrontos_excl:
            df_temp = df_temp[~df_temp['Confronto'].isin(confrontos_excl)]
        if campeonatos_excl:
            df_temp = df_temp[~df_temp['Campeonato'].isin(campeonatos_excl)]
        
        # Filtros de times (se as colunas existirem)
        if times_a_favor_excl and "Time A" in df_temp.columns and "Time B" in df_temp.columns:
            df_temp = df_temp[~((df_temp['Tip'] == df_temp['Jogador A']) & (df_temp['Time A'].isin(times_a_favor_excl))) &
                              ~((df_temp['Tip'] == df_temp['Jogador B']) & (df_temp['Time B'].isin(times_a_favor_excl)))]
        
        if times_contra_excl and "Time A" in df_temp.columns and "Time B" in df_temp.columns:
            df_temp = df_temp[~((df_temp['Tip'] != df_temp['Jogador A']) & (df_temp['Time A'].isin(times_contra_excl))) &
                              ~((df_temp['Tip'] != df_temp['Jogador B']) & (df_temp['Time B'].isin(times_contra_excl)))]
        
        # Filtros de tipo de apostas (Favorito/Azarão)
        if tipo_apostas_excl and "Favorito" in df_temp.columns and "Azarão" in df_temp.columns:
            if "Favorito" in tipo_apostas_excl:
                df_temp = df_temp[~(df_temp['Tip'] == df_temp['Favorito'])]
            if "Azarão" in tipo_apostas_excl:
                df_temp = df_temp[~(df_temp['Tip'] == df_temp['Azarão'])]
        
        # Filtros de tipo local (Mandante/Visitante)
        if tipo_local_excl:
            if "Mandante" in tipo_local_excl:
                df_temp = df_temp[~(df_temp['Tip'] == df_temp['Jogador A'])]
            if "Visitante" in tipo_local_excl:
                df_temp = df_temp[~(df_temp['Tip'] == df_temp['Jogador B'])]
        
        # Filtros de diferença de placar
        if ("Placar Envio" in df_temp.columns and 
            (diferenca_placar_min is not None or diferenca_placar_max is not None)):
            
            if "Diferença Placar" not in df_temp.columns:
                df_temp["Diferença Placar"] = df_temp["Placar Envio"].apply(self.calcular_diferenca_placar)
            
            if diferenca_placar_min is not None:
                df_temp = df_temp[df_temp["Diferença Placar"] >= diferenca_placar_min]
            
            if diferenca_placar_max is not None:
                df_temp = df_temp[df_temp["Diferença Placar"] <= diferenca_placar_max]
        
        return df_temp
    
    def iniciar_analise(self, torneio, campeonato, tip, roi_desejado_pct):
        """Iniciar análise completa com busca """
        try:
            # Converter ROI para decimal
            self.roi_desejado = float(roi_desejado_pct) / 100
            
            # Filtrar dados iniciais
            if not self.filtrar_dados_iniciais(torneio, campeonato, tip):
                return False, "Erro ao filtrar dados iniciais"
            
            # Configurar estado inicial
            self.total_inicial_apostas = len(self.df_filtrado)
            self.limite_minimo_apostas = max(1, self.total_inicial_apostas * 0.05)  # Mínimo 5% dos dados originais
            self.roi_inicial = self.calcular_roi(self.df_filtrado)
            
            if self.total_inicial_apostas == 0:
                return False, "Nenhum dado encontrado com os filtros aplicados"
            
            # Executar busca gulosa
            self.busca_gulosa()
            
            return True, f"Análise concluída! {len(self.etapas_filtros)} etapas geradas."
            
        except Exception as e:
            return False, f"Erro durante análise: {str(e)}"
    
    def busca_gulosa(self):
        """Executar busca para otimização"""
        # Configuração inicial
        self.config = {
            'w1': None, 
            'w2': None, 
            'apostas_a_favor_excl': [], 
            'apostas_contra_excl': [], 
            'confrontos': [],
            'campeonatos_excl': [],
            'times_a_favor_excl': [],
            'times_contra_excl': [],
            'tipo_apostas_excl': [],
            'tipo_local_excl': [],
            'diferenca_placar_min': None,
            'diferenca_placar_max': None
        }
        
        df_filtrado = self.df_filtrado.copy()
        roi_atual = self.calcular_roi(df_filtrado)
        self.melhor_df = df_filtrado.copy()
        self.melhor_roi = roi_atual
        self.melhor_config = self.config.copy()
        
        self.etapas_filtros = []
        
        # Etapa inicial
        self.etapas_filtros.append({
            'numero': 0,
            'ajuste': "Estado inicial",
            'entradas': len(df_filtrado),
            'lucro': df_filtrado['Lucro/Prej.'].sum(),
            'roi': roi_atual,
            'df': df_filtrado.copy(),
            'config': self.config.copy()
        })
        
        contador_etapas = 1
        max_iteracoes = 100
        iteracoes_sem_melhoria = 0
        max_sem_melhoria = 10
        
        # Progress bar para o Streamlit
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        while contador_etapas < max_iteracoes:
            status_text.text(f"🔄 Processando etapa {contador_etapas}...")
            progress_bar.progress(min(contador_etapas / max_iteracoes, 0.95))
            
            ajustes_possiveis = []
            
            # Teste de ajustes de Winrate 1
            if self.busca_config.get('usar_winrate1', True):
                winrate_1_valores = sorted(df_filtrado['Winrate 1'].dropna().unique())
                for w1 in winrate_1_valores:
                    if self.config['w1'] is None or w1 > self.config['w1']:
                        config_test = self.config.copy()
                        config_test['w1'] = w1
                        df_teste = self.aplicar_filtros(df_filtrado, config_test)
                        if len(df_teste) >= self.limite_minimo_apostas:
                            roi_teste = self.calcular_roi(df_teste)
                            if roi_teste > roi_atual:
                                impacto = roi_teste - roi_atual
                                ajustes_possiveis.append(('w1', w1, impacto, df_teste, config_test))
            
            # Teste de ajustes de Winrate 2
            if self.busca_config.get('usar_winrate2', True):
                winrate_2_valores = sorted(df_filtrado['Winrate 2'].dropna().unique())
                for w2 in winrate_2_valores:
                    if self.config['w2'] is None or w2 > self.config['w2']:
                        config_test = self.config.copy()
                        config_test['w2'] = w2
                        df_teste = self.aplicar_filtros(df_filtrado, config_test)
                        if len(df_teste) >= self.limite_minimo_apostas:
                            roi_teste = self.calcular_roi(df_teste)
                            if roi_teste > roi_atual:
                                impacto = roi_teste - roi_atual
                                ajustes_possiveis.append(('w2', w2, impacto, df_teste, config_test))
            
            # Exclusão de campeonatos prejudiciais
            if self.busca_config.get('usar_excl_campeonatos', True) and 'Campeonato' in df_filtrado.columns:
                lucro_por_campeonato = df_filtrado.groupby('Campeonato')['Lucro/Prej.'].agg(['sum', 'count']).reset_index()
                lucro_por_campeonato.columns = ['Campeonato', 'Lucro', 'Quantidade']
                campeonatos_prejudiciais = lucro_por_campeonato[
                    (lucro_por_campeonato['Lucro'] < 0) & 
                    (lucro_por_campeonato['Quantidade'] >= self.min_entradas_config.get('min_campeonatos', 3))
                ].sort_values(by='Lucro')
                
                for _, row in campeonatos_prejudiciais.iterrows():
                    campeonato = row['Campeonato']
                    if campeonato not in self.config['campeonatos_excl']:
                        config_test = self.config.copy()
                        config_test['campeonatos_excl'] = self.config['campeonatos_excl'] + [campeonato]
                        df_teste = self.aplicar_filtros(df_filtrado, config_test)
                        if len(df_teste) >= self.limite_minimo_apostas:
                            roi_teste = self.calcular_roi(df_teste)
                            if roi_teste > roi_atual:
                                impacto = roi_teste - roi_atual
                                ajustes_possiveis.append(('campeonato', campeonato, impacto, df_teste, config_test))
            
            # Exclusão de apostas a favor prejudiciais
            if self.busca_config.get('usar_excl_apostas_a_favor', True):
                lucro_a_favor = df_filtrado[df_filtrado['Tip'] == df_filtrado['Jogador A']].groupby('Jogador A')['Lucro/Prej.'].sum().reset_index()
                lucro_a_favor_b = df_filtrado[df_filtrado['Tip'] == df_filtrado['Jogador B']].groupby('Jogador B')['Lucro/Prej.'].sum().reset_index()
                lucro_a_favor = pd.concat([lucro_a_favor.rename(columns={'Jogador A': 'Jogador'}),
                                           lucro_a_favor_b.rename(columns={'Jogador B': 'Jogador'})])
                lucro_a_favor = lucro_a_favor.groupby('Jogador')['Lucro/Prej.'].sum().reset_index()
                contagem_a_favor = pd.concat([df_filtrado[df_filtrado['Tip'] == df_filtrado['Jogador A']]['Jogador A'],
                                              df_filtrado[df_filtrado['Tip'] == df_filtrado['Jogador B']]['Jogador B']]).value_counts()
                jogadores_prejudiciais_a_favor = lucro_a_favor[lucro_a_favor['Lucro/Prej.'] < 0].sort_values(by='Lucro/Prej.')
                
                for _, row in jogadores_prejudiciais_a_favor.head(5).iterrows():  # Limitar para performance
                    jogador = row['Jogador']
                    min_quantidade = self.min_entradas_config.get('min_apostas_a_favor', 3)
                    if jogador not in self.config['apostas_a_favor_excl'] and contagem_a_favor.get(jogador, 0) >= min_quantidade:
                        config_test = self.config.copy()
                        config_test['apostas_a_favor_excl'] = self.config['apostas_a_favor_excl'] + [jogador]
                        df_teste = self.aplicar_filtros(df_filtrado, config_test)
                        if len(df_teste) >= self.limite_minimo_apostas:
                            roi_teste = self.calcular_roi(df_teste)
                            if roi_teste > roi_atual:
                                impacto = roi_teste - roi_atual
                                ajustes_possiveis.append(('apostas_a_favor', jogador, impacto, df_teste, config_test))
            
            # Exclusão de apostas contra jogadores prejudiciais
            if self.busca_config.get('usar_excl_apostas_contra', True):
                lucro_contra = df_filtrado[df_filtrado['Tip'] != df_filtrado['Jogador A']].groupby('Jogador A')['Lucro/Prej.'].sum().reset_index()
                lucro_contra_b = df_filtrado[df_filtrado['Tip'] != df_filtrado['Jogador B']].groupby('Jogador B')['Lucro/Prej.'].sum().reset_index()
                lucro_contra = pd.concat([lucro_contra.rename(columns={'Jogador A': 'Jogador'}),
                                          lucro_contra_b.rename(columns={'Jogador B': 'Jogador'})])
                lucro_contra = lucro_contra.groupby('Jogador')['Lucro/Prej.'].sum().reset_index()
                contagem_contra = pd.concat([df_filtrado[df_filtrado['Tip'] != df_filtrado['Jogador A']]['Jogador A'],
                                             df_filtrado[df_filtrado['Tip'] != df_filtrado['Jogador B']]['Jogador B']]).value_counts()
                jogadores_prejudiciais_contra = lucro_contra[lucro_contra['Lucro/Prej.'] < 0].sort_values(by='Lucro/Prej.')
                
                for _, row in jogadores_prejudiciais_contra.head(5).iterrows():  # Limitar para performance
                    jogador = row['Jogador']
                    min_quantidade = self.min_entradas_config.get('min_apostas_contra', 3)
                    if jogador not in self.config['apostas_contra_excl'] and contagem_contra.get(jogador, 0) >= min_quantidade:
                        config_test = self.config.copy()
                        config_test['apostas_contra_excl'] = self.config['apostas_contra_excl'] + [jogador]
                        df_teste = self.aplicar_filtros(df_filtrado, config_test)
                        if len(df_teste) >= self.limite_minimo_apostas:
                            roi_teste = self.calcular_roi(df_teste)
                            if roi_teste > roi_atual:
                                impacto = roi_teste - roi_atual
                                ajustes_possiveis.append(('apostas_contra', jogador, impacto, df_teste, config_test))
            
            # Exclusão de confrontos prejudiciais
            if self.busca_config.get('usar_excl_confrontos', True):
                lucro_por_confronto = df_filtrado.groupby('Confronto')['Lucro/Prej.'].agg(['sum', 'count']).reset_index()
                lucro_por_confronto.columns = ['Confronto', 'Lucro', 'Quantidade']
                confrontos_prejudiciais = lucro_por_confronto[
                    (lucro_por_confronto['Lucro'] < 0) & 
                    (lucro_por_confronto['Quantidade'] >= self.min_entradas_config.get('min_confrontos', 3))
                ].sort_values(by='Lucro')
                
                for _, row in confrontos_prejudiciais.head(5).iterrows():  # Limitar para performance
                    confronto = row['Confronto']
                    if confronto not in self.config['confrontos']:
                        config_test = self.config.copy()
                        config_test['confrontos'] = self.config['confrontos'] + [confronto]
                        df_teste = self.aplicar_filtros(df_filtrado, config_test)
                        if len(df_teste) >= self.limite_minimo_apostas:
                            roi_teste = self.calcular_roi(df_teste)
                            if roi_teste > roi_atual:
                                impacto = roi_teste - roi_atual
                                ajustes_possiveis.append(('confronto', confronto, impacto, df_teste, config_test))
            
            # Exclusão de apostas por tipo (Favorito/Azarão) 
            if self.busca_config.get('usar_excl_tipo_apostas', True) and "Favorito" in df_filtrado.columns and "Azarão" in df_filtrado.columns:
                tipos_disponiveis = ["Favorito", "Azarão"]
                tipos_nao_excluidos = [tipo for tipo in tipos_disponiveis if tipo not in self.config['tipo_apostas_excl']]
                
                for tipo in tipos_nao_excluidos:
                    if tipo == "Favorito":
                        quantidade_apostas = df_filtrado[df_filtrado['Tip'] == df_filtrado['Favorito']].shape[0]
                    else:
                        quantidade_apostas = df_filtrado[df_filtrado['Tip'] == df_filtrado['Azarão']].shape[0]
                    
                    min_quantidade = self.min_entradas_config.get('min_tipo_apostas', 3)
                    if quantidade_apostas >= min_quantidade:
                        config_test = self.config.copy()
                        config_test['tipo_apostas_excl'] = self.config['tipo_apostas_excl'] + [tipo]
                        df_teste = self.aplicar_filtros(df_filtrado, config_test)
                        if len(df_teste) >= self.limite_minimo_apostas:
                            roi_teste = self.calcular_roi(df_teste)
                            if roi_teste > roi_atual:
                                impacto = roi_teste - roi_atual
                                ajustes_possiveis.append(('tipo_aposta', tipo, impacto, df_teste, config_test))
            
            # Exclusão de apostas por tipo local (Mandante/Visitante)
            if self.busca_config.get('usar_excl_tipo_local', True):
                tipos_local_disponiveis = ["Mandante", "Visitante"]
                tipos_local_nao_excluidos = [tipo for tipo in tipos_local_disponiveis if tipo not in self.config['tipo_local_excl']]
                
                for tipo_local in tipos_local_nao_excluidos:
                    if tipo_local == "Mandante":
                        quantidade_apostas = df_filtrado[df_filtrado['Tip'] == df_filtrado['Jogador A']].shape[0]
                    else:
                        quantidade_apostas = df_filtrado[df_filtrado['Tip'] == df_filtrado['Jogador B']].shape[0]
                    
                    min_quantidade = self.min_entradas_config.get('min_tipo_local', 3)
                    if quantidade_apostas >= min_quantidade:
                        config_test = self.config.copy()
                        config_test['tipo_local_excl'] = self.config['tipo_local_excl'] + [tipo_local]
                        df_teste = self.aplicar_filtros(df_filtrado, config_test)
                        if len(df_teste) >= self.limite_minimo_apostas:
                            roi_teste = self.calcular_roi(df_teste)
                            if roi_teste > roi_atual:
                                impacto = roi_teste - roi_atual
                                ajustes_possiveis.append(('tipo_local', tipo_local, impacto, df_teste, config_test))
            
            # Exclusão de times a favor
            if self.busca_config.get('usar_excl_times_a_favor', True) and "Time A" in df_filtrado.columns and "Time B" in df_filtrado.columns:
                times_a_favor_df = df_filtrado.copy()
                times_a_favor_df['Time Escolhido'] = times_a_favor_df.apply(
                    lambda row: row['Time A'] if row['Tip'] == row['Jogador A'] else (
                        row['Time B'] if row['Tip'] == row['Jogador B'] else None
                    ), axis=1
                )
                
                lucro_por_time = times_a_favor_df.groupby('Time Escolhido')['Lucro/Prej.'].agg(['sum', 'count']).reset_index()
                lucro_por_time.columns = ['Time', 'Lucro', 'Quantidade']
                min_quantidade = self.min_entradas_config.get('min_times_a_favor', 3)
                times_prejudiciais = lucro_por_time[(lucro_por_time['Lucro'] < 0) & (lucro_por_time['Quantidade'] >= min_quantidade)].sort_values(by='Lucro')
                
                for _, row in times_prejudiciais.head(3).iterrows():  # Limitar para performance
                    time = row['Time']
                    if time not in self.config['times_a_favor_excl'] and pd.notna(time):
                        config_test = self.config.copy()
                        config_test['times_a_favor_excl'] = self.config['times_a_favor_excl'] + [time]
                        df_teste = self.aplicar_filtros(df_filtrado, config_test)
                        if len(df_teste) >= self.limite_minimo_apostas:
                            roi_teste = self.calcular_roi(df_teste)
                            if roi_teste > roi_atual:
                                impacto = roi_teste - roi_atual
                                ajustes_possiveis.append(('time_a_favor', time, impacto, df_teste, config_test))
            
            # Exclusão de times contra
            if self.busca_config.get('usar_excl_times_contra', True) and "Time A" in df_filtrado.columns and "Time B" in df_filtrado.columns:
                times_contra_df = df_filtrado.copy()
                times_contra_df['Time Contra'] = times_contra_df.apply(
                    lambda row: row['Time B'] if row['Tip'] == row['Jogador A'] else (
                        row['Time A'] if row['Tip'] == row['Jogador B'] else None
                    ), axis=1
                )
                
                lucro_por_time_contra = times_contra_df.groupby('Time Contra')['Lucro/Prej.'].agg(['sum', 'count']).reset_index()
                lucro_por_time_contra.columns = ['Time', 'Lucro', 'Quantidade']
                min_quantidade = self.min_entradas_config.get('min_times_contra', 3)
                times_contra_prejudiciais = lucro_por_time_contra[(lucro_por_time_contra['Lucro'] < 0) & (lucro_por_time_contra['Quantidade'] >= min_quantidade)].sort_values(by='Lucro')
                
                for _, row in times_contra_prejudiciais.head(3).iterrows():  # Limitar para performance
                    time = row['Time']
                    if time not in self.config['times_contra_excl'] and pd.notna(time):
                        config_test = self.config.copy()
                        config_test['times_contra_excl'] = self.config['times_contra_excl'] + [time]
                        df_teste = self.aplicar_filtros(df_filtrado, config_test)
                        if len(df_teste) >= self.limite_minimo_apostas:
                            roi_teste = self.calcular_roi(df_teste)
                            if roi_teste > roi_atual:
                                impacto = roi_teste - roi_atual
                                ajustes_possiveis.append(('time_contra', time, impacto, df_teste, config_test))
            
            # Ajuste de diferença de placar mínima
            if self.busca_config.get('usar_diferenca_placar_min', True) and "Placar Envio" in df_filtrado.columns:
                if "Diferença Placar" not in df_filtrado.columns:
                    df_filtrado["Diferença Placar"] = df_filtrado["Placar Envio"].apply(self.calcular_diferenca_placar)
                
                diferencas_disponiveis = sorted(df_filtrado["Diferença Placar"].dropna().unique())
                
                for dif_min in diferencas_disponiveis[:5]:  # Limitar para performance
                    if self.config['diferenca_placar_min'] is None or dif_min > self.config['diferenca_placar_min']:
                        config_test = self.config.copy()
                        config_test['diferenca_placar_min'] = dif_min
                        df_teste = self.aplicar_filtros(df_filtrado, config_test)
                        if len(df_teste) >= self.limite_minimo_apostas:
                            roi_teste = self.calcular_roi(df_teste)
                            if roi_teste > roi_atual:
                                impacto = roi_teste - roi_atual
                                ajustes_possiveis.append(('diferenca_placar_min', dif_min, impacto, df_teste, config_test))
            
            # Ajuste de diferença de placar máxima
            if self.busca_config.get('usar_diferenca_placar_max', True) and "Placar Envio" in df_filtrado.columns:
                if "Diferença Placar" not in df_filtrado.columns:
                    df_filtrado["Diferença Placar"] = df_filtrado["Placar Envio"].apply(self.calcular_diferenca_placar)
                
                diferencas_disponiveis = sorted(df_filtrado["Diferença Placar"].dropna().unique(), reverse=True)
                
                for dif_max in diferencas_disponiveis[:5]:  # Limitar para performance
                    if self.config['diferenca_placar_max'] is None or dif_max < self.config['diferenca_placar_max']:
                        config_test = self.config.copy()
                        config_test['diferenca_placar_max'] = dif_max
                        df_teste = self.aplicar_filtros(df_filtrado, config_test)
                        if len(df_teste) >= self.limite_minimo_apostas:
                            roi_teste = self.calcular_roi(df_teste)
                            if roi_teste > roi_atual:
                                impacto = roi_teste - roi_atual
                                ajustes_possiveis.append(('diferenca_placar_max', dif_max, impacto, df_teste, config_test))
            
            # Se não há ajustes possíveis, parar
            if not ajustes_possiveis:
                iteracoes_sem_melhoria += 1
                if iteracoes_sem_melhoria >= max_sem_melhoria:
                    break
                contador_etapas += 1
                continue
            
            # Ordenar por impacto e escolher o melhor
            ajustes_possiveis.sort(key=lambda x: x[2], reverse=True)
            melhor_ajuste = ajustes_possiveis[0]
            tipo, valor, _, df_novo, config_novo = melhor_ajuste
            
            # Criar descrição do ajuste
            if tipo == 'w1':
                descricao_ajuste = f"Winrate 1 mínimo = {valor:.2f}%"
            elif tipo == 'w2':
                descricao_ajuste = f"Winrate 2 mínimo = {valor:.2f}%"
            elif tipo == 'apostas_a_favor':
                descricao_ajuste = f"Excluídas apostas a favor de {valor}"
            elif tipo == 'apostas_contra':
                descricao_ajuste = f"Excluídas apostas contra {valor}"
            elif tipo == 'confronto':
                descricao_ajuste = f"Excluído confronto {valor}"
            elif tipo == 'campeonato':
                descricao_ajuste = f"Excluído campeonato {valor}"
            elif tipo == 'tipo_aposta':
                descricao_ajuste = f"Excluídas apostas a favor do {valor}"
            elif tipo == 'tipo_local':
                descricao_ajuste = f"Excluídas apostas a favor do {valor}"
            elif tipo == 'time_a_favor':
                descricao_ajuste = f"Excluídas apostas a favor do time {valor}"
            elif tipo == 'time_contra':
                descricao_ajuste = f"Excluídas apostas contra o time {valor}"
            elif tipo == 'diferenca_placar_min':
                descricao_ajuste = f"Diferença de placar mínima = {valor}"
            elif tipo == 'diferenca_placar_max':
                descricao_ajuste = f"Diferença de placar máxima = {valor}"
            else:
                descricao_ajuste = f"Aplicado filtro {tipo}: {valor}"
            
            # Atualizar estado
            df_filtrado = df_novo.copy()
            self.config = config_novo
            roi_atual = self.calcular_roi(df_filtrado)
            lucro_atual = df_filtrado['Lucro/Prej.'].sum()
            
            # Adicionar etapa
            self.etapas_filtros.append({
                'numero': contador_etapas,
                'ajuste': descricao_ajuste,
                'entradas': len(df_filtrado),
                'lucro': lucro_atual,
                'roi': roi_atual,
                'df': df_filtrado.copy(),
                'config': self.config.copy()
            })
            
            # Atualizar melhor resultado se necessário
            if roi_atual > self.melhor_roi:
                self.melhor_df = df_filtrado.copy()
                self.melhor_roi = roi_atual
                self.melhor_config = self.config.copy()
                iteracoes_sem_melhoria = 0
            else:
                iteracoes_sem_melhoria += 1
            
            # Verificar se atingiu o limite mínimo de apostas
            if len(df_filtrado) < self.limite_minimo_apostas:
                break
            
            contador_etapas += 1
        
        progress_bar.progress(1.0)
        status_text.text("✅ Análise concluída!")
    
    def gerar_relatorio_excel(self, etapa_numero):
        """Gerar relatório Excel para uma etapa específica"""
        if not self.etapas_filtros or etapa_numero >= len(self.etapas_filtros):
            return None, None
        
        etapa = self.etapas_filtros[etapa_numero]
        df_final = etapa['df'].copy()
        config_final = etapa['config']
        
        # Adicionar campos normalizados para análise
        df_final["Confronto Normalizado"] = df_final.apply(
            lambda row: " vs ".join(sorted([str(row["Jogador A"]), str(row["Jogador B"])])) 
            if pd.notna(row["Jogador A"]) and pd.notna(row["Jogador B"]) else "", axis=1
        )

        # Adicionar coluna Jogador Contra
        df_final['Jogador Contra'] = df_final.apply(
            lambda row: row['Jogador B'] if row['Tip'] == row['Jogador A'] else (
                row['Jogador A'] if row['Tip'] == row['Jogador B'] else None
            ), axis=1
        )

        # Adicionar colunas Time a Favor e Time Contra (se as colunas Time A e Time B existirem)
        if "Time A" in df_final.columns and "Time B" in df_final.columns:
            df_final['Time a Favor'] = df_final.apply(
                lambda row: row['Time A'] if row['Tip'] == row['Jogador A'] else (
                    row['Time B'] if row['Tip'] == row['Jogador B'] else None
                ), axis=1
            )
            
            df_final['Time Contra'] = df_final.apply(
                lambda row: row['Time B'] if row['Tip'] == row['Jogador A'] else (
                    row['Time A'] if row['Tip'] == row['Jogador B'] else None
                ), axis=1
            )
        else:
            df_final['Time a Favor'] = None
            df_final['Time Contra'] = None

        # Adicionar coluna Diferença Placar (se a coluna Placar Envio existir)
        if "Placar Envio" in df_final.columns:
            df_final["Diferença Placar"] = df_final["Placar Envio"].apply(self.calcular_diferenca_placar)
        else:
            df_final["Diferença Placar"] = None

        # Adicionar coluna ROI (será substituída por fórmula no Excel)
        df_final["ROI"] = df_final["Lucro/Prej."]

        if "Favorito" in df_final.columns and "Azarão" in df_final.columns:
            df_final["Aposta Favor (Favorito/Azarão)"] = df_final.apply(
                lambda row: "Favorito" if pd.notna(row["Favorito"]) and row["Tip"] == row["Favorito"] 
                            else ("Azarão" if pd.notna(row["Azarão"]) and row["Tip"] == row["Azarão"] else "N/A"), axis=1
            )
        else:
            df_final["Aposta Favor (Favorito/Azarão)"] = "N/A"

        # Análises agrupadas
        df_confronto = df_final.groupby(["Torneio", "Confronto Normalizado"]).agg(
            Quantidade_Entradas=("Lucro/Prej.", "count"), 
            Lucro_Prej=("Lucro/Prej.", "sum")
        ).reset_index()
        df_confronto["ROI (%)"] = (df_confronto["Lucro_Prej"] / df_confronto["Quantidade_Entradas"]).round(4)

        if "Campeonato" in df_final.columns:
            df_campeonato = df_final.groupby(["Torneio", "Campeonato"]).agg(
                Quantidade_Entradas=("Lucro/Prej.", "count"), 
                Lucro_Prej=("Lucro/Prej.", "sum")
            ).reset_index()
            df_campeonato["ROI (%)"] = (df_campeonato["Lucro_Prej"] / df_campeonato["Quantidade_Entradas"]).round(4)
        else:
            df_campeonato = pd.DataFrame(columns=["Torneio", "Campeonato", "Quantidade_Entradas", "Lucro_Prej", "ROI (%)"])

        df_winrate1 = df_final.groupby(["Torneio", "Winrate 1"]).agg(
            Quantidade_Entradas=("Lucro/Prej.", "count"), 
            Lucro_Prej=("Lucro/Prej.", "sum")
        ).reset_index()
        df_winrate1["ROI (%)"] = (df_winrate1["Lucro_Prej"] / df_winrate1["Quantidade_Entradas"]).round(4)

        df_winrate2 = df_final.groupby(["Torneio", "Winrate 2"]).agg(
            Quantidade_Entradas=("Lucro/Prej.", "count"), 
            Lucro_Prej=("Lucro/Prej.", "sum")
        ).reset_index()
        df_winrate2["ROI (%)"] = (df_winrate2["Lucro_Prej"] / df_winrate2["Quantidade_Entradas"]).round(4)

        df_jogadores_a_favor = df_final.copy()
        df_jogadores_a_favor['Jogador Escolhido'] = df_jogadores_a_favor.apply(
            lambda row: row['Jogador A'] if row['Tip'] == row['Jogador A'] else (
                row['Jogador B'] if row['Tip'] == row['Jogador B'] else None
            ), axis=1
        )
        df_jogadores_a_favor = df_jogadores_a_favor[df_jogadores_a_favor['Jogador Escolhido'].notna()]
        df_jogador = df_jogadores_a_favor.groupby(["Torneio", "Jogador Escolhido"]).agg(
            Quantidade_Entradas=("Lucro/Prej.", "count"), 
            Lucro_Prej=("Lucro/Prej.", "sum")
        ).reset_index()
        df_jogador.rename(columns={"Jogador Escolhido": "Jogador"}, inplace=True)
        df_jogador["ROI (%)"] = (df_jogador["Lucro_Prej"] / df_jogador["Quantidade_Entradas"]).round(4)

        df_jogadores_contra = df_final.copy()
        df_jogadores_contra['Jogador Contra'] = df_jogadores_contra.apply(
            lambda row: row['Jogador B'] if row['Tip'] == row['Jogador A'] else (
                row['Jogador A'] if row['Tip'] == row['Jogador B'] else None
            ), axis=1
        )
        df_jogadores_contra = df_jogadores_contra[df_jogadores_contra['Jogador Contra'].notna()]
        df_jogador_contra = df_jogadores_contra.groupby(["Torneio", "Jogador Contra"]).agg(
            Quantidade_Entradas=("Lucro/Prej.", "count"), 
            Lucro_Prej=("Lucro/Prej.", "sum")
        ).reset_index()
        df_jogador_contra.rename(columns={"Jogador Contra": "Jogador"}, inplace=True)
        df_jogador_contra["ROI (%)"] = (df_jogador_contra["Lucro_Prej"] / df_jogador_contra["Quantidade_Entradas"]).round(4)

        df_times_a_favor = df_final.copy()
        df_times_a_favor['Time Escolhido'] = df_times_a_favor.apply(
            lambda row: row['Time A'] if row['Tip'] == row['Jogador A'] else (
                row['Time B'] if row['Tip'] == row['Jogador B'] else None
            ), axis=1
        )
        df_times_a_favor = df_times_a_favor[df_times_a_favor['Time Escolhido'].notna()]
        df_time = df_times_a_favor.groupby(["Torneio", "Time Escolhido"]).agg(
            Quantidade_Entradas=("Lucro/Prej.", "count"), 
            Lucro_Prej=("Lucro/Prej.", "sum")
        ).reset_index()
        df_time.rename(columns={"Time Escolhido": "Time"}, inplace=True)
        df_time["ROI (%)"] = (df_time["Lucro_Prej"] / df_time["Quantidade_Entradas"]).round(4)

        df_times_contra = df_final.copy()
        df_times_contra['Time Contra'] = df_times_contra.apply(
            lambda row: row['Time B'] if row['Tip'] == row['Jogador A'] else (
                row['Time A'] if row['Tip'] == row['Jogador B'] else None
            ), axis=1
        )
        df_times_contra = df_times_contra[df_times_contra['Time Contra'].notna()]
        df_time_contra = df_times_contra.groupby(["Torneio", "Time Contra"]).agg(
            Quantidade_Entradas=("Lucro/Prej.", "count"), 
            Lucro_Prej=("Lucro/Prej.", "sum")
        ).reset_index()
        df_time_contra.rename(columns={"Time Contra": "Time"}, inplace=True)
        df_time_contra["ROI (%)"] = (df_time_contra["Lucro_Prej"] / df_time_contra["Quantidade_Entradas"]).round(4)

        if "Favorito" in df_final.columns and "Azarão" in df_final.columns:
            df_apostas_favorito = df_final[df_final['Tip'] == df_final['Favorito']].copy()
            df_favorito = df_apostas_favorito.groupby(["Torneio", "Favorito"]).agg(
                Quantidade_Entradas=("Lucro/Prej.", "count"), 
                Lucro_Prej=("Lucro/Prej.", "sum")
            ).reset_index()
            df_favorito["ROI (%)"] = (df_favorito["Lucro_Prej"] / df_favorito["Quantidade_Entradas"]).round(4)
            
            df_apostas_azarao = df_final[df_final['Tip'] == df_final['Azarão']].copy()
            df_azarao = df_apostas_azarao.groupby(["Torneio", "Azarão"]).agg(
                Quantidade_Entradas=("Lucro/Prej.", "count"), 
                Lucro_Prej=("Lucro/Prej.", "sum")
            ).reset_index()
            df_azarao["ROI (%)"] = (df_azarao["Lucro_Prej"] / df_azarao["Quantidade_Entradas"]).round(4)
            
            df_tipo_aposta = df_final.copy()
            df_tipo_aposta['Tipo Aposta'] = df_tipo_aposta.apply(
                lambda row: 'Favorito' if row['Tip'] == row['Favorito'] else (
                    'Azarão' if row['Tip'] == row['Azarão'] else 'Não Classificado'
                ), axis=1
            )
            df_tipo = df_tipo_aposta.groupby(["Torneio", "Tipo Aposta"]).agg(
                Quantidade_Entradas=("Lucro/Prej.", "count"), 
                Lucro_Prej=("Lucro/Prej.", "sum")
            ).reset_index()
            df_tipo["ROI (%)"] = (df_tipo["Lucro_Prej"] / df_tipo["Quantidade_Entradas"]).round(4)

        df_tipo_local = df_final.copy()
        df_tipo_local['Tipo Local'] = df_tipo_local.apply(
            lambda row: 'Mandante' if row['Tip'] == row['Jogador A'] else (
                'Visitante' if row['Tip'] == row['Jogador B'] else 'Não Classificado'
            ), axis=1
        )
        df_local = df_tipo_local.groupby(["Torneio", "Tipo Local"]).agg(
            Quantidade_Entradas=("Lucro/Prej.", "count"), 
            Lucro_Prej=("Lucro/Prej.", "sum")
        ).reset_index()
        df_local["ROI (%)"] = (df_local["Lucro_Prej"] / df_local["Quantidade_Entradas"]).round(4)

        df_final['Tipo Local'] = df_final.apply(
            lambda row: 'Mandante' if row['Tip'] == row['Jogador A'] else (
                'Visitante' if row['Tip'] == row['Jogador B'] else 'Não Classificado'
            ), axis=1
        )

        df_linha = df_final.groupby(["Torneio", "Linha"]).agg(
            Quantidade_Entradas=("Lucro/Prej.", "count"), 
            Lucro_Prej=("Lucro/Prej.", "sum")
        ).reset_index()
        df_linha["ROI (%)"] = (df_linha["Lucro_Prej"] / df_linha["Quantidade_Entradas"]).round(4)

        if "Placar Envio" in df_final.columns:
            df_placar_envio = df_final.groupby(["Torneio", "Placar Envio"]).agg(
                Quantidade_Entradas=("Lucro/Prej.", "count"), 
                Lucro_Prej=("Lucro/Prej.", "sum")
            ).reset_index()
            df_placar_envio["ROI (%)"] = (df_placar_envio["Lucro_Prej"] / df_placar_envio["Quantidade_Entradas"]).round(4)
            
            df_diferenca_placar = df_final.dropna(subset=["Diferença Placar"]).groupby(["Torneio", "Diferença Placar"]).agg(
                Quantidade_Entradas=("Lucro/Prej.", "count"), 
                Lucro_Prej=("Lucro/Prej.", "sum")
            ).reset_index()
            df_diferenca_placar["ROI (%)"] = (df_diferenca_placar["Lucro_Prej"] / df_diferenca_placar["Quantidade_Entradas"]).round(4)
            df_diferenca_placar = df_diferenca_placar.sort_values("Diferença Placar")

        # Criar arquivo Excel em memória
        output = io.BytesIO()
        
        # Salvar em Excel com várias planilhas
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df_final.to_excel(writer, sheet_name="Tips Enviadas", index=False)
            df_confronto.to_excel(writer, sheet_name="Confronto", index=False)
            
            if "Campeonato" in df_final.columns:
                df_campeonato.to_excel(writer, sheet_name="Campeonato", index=False)
            
            df_winrate1.to_excel(writer, sheet_name="Winrate 1", index=False)
            df_winrate2.to_excel(writer, sheet_name="Winrate 2", index=False)
            df_jogador.to_excel(writer, sheet_name="Jogador", index=False)
            df_jogador_contra.to_excel(writer, sheet_name="Jogador Contra", index=False)
            
            if "Time A" in df_final.columns and "Time B" in df_final.columns:
                df_time.to_excel(writer, sheet_name="Time", index=False)
                df_time_contra.to_excel(writer, sheet_name="Time Contra", index=False)
            
            df_linha.to_excel(writer, sheet_name="Linha", index=False)
            
            if "Placar Envio" in df_final.columns:
                df_placar_envio.to_excel(writer, sheet_name="Placar Envio", index=False)
                df_diferenca_placar.to_excel(writer, sheet_name="Diferença Placar", index=False)
            
            if "Favorito" in df_final.columns and "Azarão" in df_final.columns:
                df_favorito.to_excel(writer, sheet_name="Jogador Favorito", index=False)
                df_azarao.to_excel(writer, sheet_name="Jogador Azarão", index=False)
                df_tipo.to_excel(writer, sheet_name="Tipo Aposta", index=False)
            
            df_local.to_excel(writer, sheet_name="Tipo Local", index=False)

        # Aplicar formatação (similar ao original)
        excel_data = self._aplicar_formatacao_excel(output, etapa_numero)
        
        # Gerar arquivo TXT de configuração
        config_texto = self._gerar_config_texto(etapa_numero, etapa, config_final)
        
        return excel_data, config_texto
    
    def _aplicar_formatacao_excel(self, output, etapa_numero):
        """Aplicar formatação completa ao Excel como no original"""
        try:
            # Recarregar o workbook para aplicar formatações
            wb = load_workbook(output)
            
            # Lista de sheets para formatar
            sheets_to_format = ["Tips Enviadas", "Campeonato", "Confronto", "Winrate 1", "Winrate 2", 
                              "Jogador", "Jogador Contra", "Jogador Favorito", "Jogador Azarão", 
                              "Tipo Aposta", "Tipo Local", "Linha", "Time", "Time Contra", 
                              "Placar Envio", "Diferença Placar"]
            
            sheets_to_format_lucro = ["Campeonato", "Confronto", "Winrate 1", "Winrate 2", "Jogador", 
                                    "Jogador Contra", "Jogador Favorito", "Jogador Azarão", "Tipo Aposta", 
                                    "Tipo Local", "Linha", "Time", "Time Contra", "Placar Envio", "Diferença Placar"]
            
            sheets_to_format_roi = ["Campeonato", "Confronto", "Winrate 1", "Winrate 2", "Jogador", 
                                  "Jogador Contra", "Jogador Favorito", "Jogador Azarão", "Tipo Aposta", 
                                  "Tipo Local", "Linha", "Time", "Time Contra", "Placar Envio", "Diferença Placar"]
            
            table_style = "TableStyleLight1"

            for sheet_name in sheets_to_format:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    if ws.max_row > 1 and ws.max_column > 1:
                        # Criar tabela
                        table_range = f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"
                        table_name = f"Table_{sheet_name.replace(' ', '_')}"
                        
                        # Verificar se já existe tabela com esse nome
                        existing_tables = [t.name for t in ws.tables.values()]
                        if table_name in existing_tables:
                            table_name = f"{table_name}_{len(existing_tables)}"
                        
                        table = Table(displayName=table_name, ref=table_range)
                        style = TableStyleInfo(name=table_style, showFirstColumn=False, 
                                             showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                        table.tableStyleInfo = style
                        ws.add_table(table)

                        # Formatação do cabeçalho
                        header_fill = PatternFill(start_color="B2B2B2", end_color="B2B2B2", fill_type="solid")
                        header_font = Font(color="FFFFFF", bold=True)
                        for cell in ws[1]:
                            cell.fill = header_fill
                            cell.font = header_font

                        # Ajustar largura das colunas
                        for col in range(1, ws.max_column + 1):
                            column_letter = get_column_letter(col)
                            max_length = 0
                            for cell in ws[column_letter]:
                                try:
                                    cell_value = str(cell.value) if cell.value is not None else ""
                                    max_length = max(max_length, len(cell_value))
                                except (TypeError, ValueError):
                                    continue
                            ws.column_dimensions[column_letter].width = max_length * 1.2 if max_length > 0 else 10

                        # Cores para valores positivos e negativos
                        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        green_font = Font(color="006400", bold=True)
                        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        red_font = Font(color="8B0000", bold=True)

                        # Encontrar colunas de lucro e ROI
                        lucro_col_idx = None
                        roi_col_idx = None
                        roi_formula_col_idx = None
                        resultado_col_idx = None
                        
                        for col_idx, cell in enumerate(ws[1], start=1):
                            if cell.value in ["Lucro_Prej", "Lucro/Prej."]:
                                lucro_col_idx = col_idx
                            elif cell.value == "ROI (%)":
                                roi_col_idx = col_idx
                            elif cell.value == "ROI":
                                roi_formula_col_idx = col_idx
                            elif cell.value == "Resultado":
                                resultado_col_idx = col_idx

                        # Formatação da coluna de lucro
                        if sheet_name in sheets_to_format_lucro and lucro_col_idx:
                            for row in ws.iter_rows(min_row=2, min_col=lucro_col_idx, max_col=lucro_col_idx):
                                for cell in row:
                                    if pd.notna(cell.value):
                                        cell.number_format = '0.00'
                                        try:
                                            value = float(cell.value)
                                            if value > 0:
                                                cell.fill = green_fill
                                                cell.font = green_font
                                            elif value < 0:
                                                cell.fill = red_fill
                                                cell.font = red_font
                                        except (ValueError, TypeError):
                                            pass

                        # Formatação da coluna de ROI
                        if sheet_name in sheets_to_format_roi and roi_col_idx:
                            for row in ws.iter_rows(min_row=2, min_col=roi_col_idx, max_col=roi_col_idx):
                                for cell in row:
                                    if pd.notna(cell.value):
                                        cell.number_format = '0.0%'
                                        try:
                                            value = float(cell.value)
                                            if value > 0:
                                                cell.fill = green_fill
                                                cell.font = green_font
                                            elif value < 0:
                                                cell.fill = red_fill
                                                cell.font = red_font
                                        except (ValueError, TypeError):
                                            pass
                        
                        # Formatação da coluna Resultado (Green/Red)
                        if sheet_name == "Tips Enviadas" and resultado_col_idx:
                            for row in ws.iter_rows(min_row=2, min_col=resultado_col_idx, max_col=resultado_col_idx):
                                for cell in row:
                                    if pd.notna(cell.value):
                                        valor_texto = str(cell.value).lower()
                                        if "green" in valor_texto:
                                            cell.fill = green_fill
                                            cell.font = green_font
                                        elif "red" in valor_texto:
                                            cell.fill = red_fill
                                            cell.font = red_font

                        # Adicionar fórmulas na coluna ROI da planilha Tips Enviadas
                        if sheet_name == "Tips Enviadas" and roi_formula_col_idx and lucro_col_idx:
                            lucro_col_letter = get_column_letter(lucro_col_idx)
                            for row_num in range(2, ws.max_row + 1):
                                cell = ws.cell(row=row_num, column=roi_formula_col_idx)
                                cell.value = f"={lucro_col_letter}{row_num}/1"
                                cell.number_format = '0.00%'

            # Salvar as alterações
            output_formatted = io.BytesIO()
            wb.save(output_formatted)
            return output_formatted.getvalue()
            
        except Exception as e:
            # Se houver erro na formatação, retornar arquivo original
            return output.getvalue()
    
    def _gerar_config_texto(self, etapa_numero, etapa, config_final):
        """Gerar texto de configuração"""
        config_table = [
            ["Etapa Escolhida", etapa_numero],
            ["Total de Apostas", etapa['entradas']],
            ["ROI", f"{etapa['roi']:.3f}"],
            ["Lucro Total", f"{etapa['lucro']:.2f}"],
            ["Winrate 1 Mínimo", f"{config_final['w1']:.2f}%" if config_final['w1'] else "Nenhum"],
            ["Winrate 2 Mínimo", f"{config_final['w2']:.2f}%" if config_final['w2'] else "Nenhum"],
            ["Apostas a Favor Excluídas", ", ".join(config_final['apostas_a_favor_excl']) if config_final['apostas_a_favor_excl'] else "Nenhum"],
            ["Apostas Contra Excluídas", ", ".join(config_final['apostas_contra_excl']) if config_final['apostas_contra_excl'] else "Nenhum"],
            ["Confrontos Excluídos", ", ".join(config_final['confrontos']) if config_final['confrontos'] else "Nenhum"],
            ["Campeonatos Excluídos", ", ".join(config_final['campeonatos_excl']) if config_final['campeonatos_excl'] else "Nenhum"],
            ["Tipos de Aposta Excluídos", ", ".join(config_final['tipo_apostas_excl']) if config_final['tipo_apostas_excl'] else "Nenhum"],
            ["Tipos de Local Excluídos", ", ".join(config_final['tipo_local_excl']) if config_final['tipo_local_excl'] else "Nenhum"],
            ["Times a Favor Excluídos", ", ".join(config_final['times_a_favor_excl']) if config_final['times_a_favor_excl'] else "Nenhum"],
            ["Times Contra Excluídos", ", ".join(config_final['times_contra_excl']) if config_final['times_contra_excl'] else "Nenhum"],
            ["Diferença de Placar Mínima", str(config_final['diferenca_placar_min']) if config_final['diferenca_placar_min'] is not None else "Nenhum"],
            ["Diferença de Placar Máxima", str(config_final['diferenca_placar_max']) if config_final['diferenca_placar_max'] is not None else "Nenhum"]
        ]
        
        texto = "Configuração da Etapa Escolhida:\n"
        texto += tabulate(config_table, headers=["Parâmetro", "Valor"], tablefmt="pretty")
        texto += "\n\n"
        
        return texto

# Função para criar link de download
def get_download_link(data, filename, text):
    """Gera link de download para dados binários"""
    b64 = base64.b64encode(data).decode()
    href = f'<a href="data:application/octet-stream;base64,{b64}" download="{filename}">{text}</a>'
    return href

# Interface Streamlit
def main():
    # Header principal
    st.markdown("""
    <div class="main-header">
        <h1>🏆 Handicap/ML Pro</h1>
        <p>Análise avançada com busca otimizada para maximizar seu ROI no TipManager</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Inicializar analyzer na sessão
    if 'analyzer' not in st.session_state:
        st.session_state.analyzer = BacktestAnalyzer()
    
    analyzer = st.session_state.analyzer
    
    # Sidebar para configurações
    with st.sidebar:
        st.markdown('<div class="section-header">📁 Arquivo</div>', unsafe_allow_html=True)
        
        uploaded_file = st.file_uploader(
            "📤 Envie um arquivo Excel (.xlsx/.xls)",
            type=['xlsx', 'xls'],
            help="Faça upload da planilha de dados para análise"
        )
        
        if uploaded_file is not None:
            if analyzer.carregar_arquivo(uploaded_file):
                st.success("✅ Arquivo carregado com sucesso!")
                st.session_state['file_uploaded'] = True
            else:
                st.session_state['file_uploaded'] = False
        
        st.markdown("---")
        
        # Configurações da busca gulosa
        st.markdown('<div class="section-header">⚙️ Busca </div>', unsafe_allow_html=True)
        
        # Criar abas para organizar melhor
        tab1, tab2 = st.tabs(["🎯 Básico", "🔧 Busca "])
        
        with tab2:
            st.markdown("""
            <div class="config-section">
                <div class="config-title">🔧 Configurações da Busca </div>
                <p style="color: #6b7280; font-size: 0.9rem;">Selecione quais tipos de filtros serão aplicados durante a otimização</p>
            </div>
            """, unsafe_allow_html=True)
            
            # Filtros de Winrate
            st.markdown("### 📊 Filtros de Winrate")
            col1, col2 = st.columns(2)
            with col1:
                analyzer.busca_config['usar_winrate1'] = st.checkbox(
                    "Ajustar Winrate 1 mínimo", 
                    value=analyzer.busca_config.get('usar_winrate1', True),
                    help="Permite ajustar o winrate mínimo do primeiro jogador"
                )
                if 'min_winrate1' not in analyzer.min_entradas_config:
                    analyzer.min_entradas_config['min_winrate1'] = 10
                analyzer.min_entradas_config['min_winrate1'] = st.number_input(
                    "Min. entradas:", min_value=1, max_value=50, 
                    value=analyzer.min_entradas_config.get('min_winrate1', 10),
                    key="min_winrate1"
                )
            with col2:
                analyzer.busca_config['usar_winrate2'] = st.checkbox(
                    "Ajustar Winrate 2 mínimo", 
                    value=analyzer.busca_config.get('usar_winrate2', True),
                    help="Permite ajustar o winrate mínimo do segundo jogador"
                )
                if 'min_winrate2' not in analyzer.min_entradas_config:
                    analyzer.min_entradas_config['min_winrate2'] = 10
                analyzer.min_entradas_config['min_winrate2'] = st.number_input(
                    "Min. entradas:", min_value=1, max_value=50, 
                    value=analyzer.min_entradas_config.get('min_winrate2', 10),
                    key="min_winrate2"
                )
            
            st.markdown("---")
            
            # Filtros de Exclusão
            st.markdown("### 🚫 Filtros de Exclusão")
            
            # Primeira linha de exclusões
            col1, col2 = st.columns(2)
            with col1:
                analyzer.busca_config['usar_excl_campeonatos'] = st.checkbox(
                    "Excluir campeonatos prejudiciais", 
                    value=analyzer.busca_config.get('usar_excl_campeonatos', True),
                    help="Exclui campeonatos com ROI negativo"
                )
                analyzer.min_entradas_config['min_campeonatos'] = st.number_input(
                    "Min. entradas:", min_value=1, max_value=50, 
                    value=analyzer.min_entradas_config.get('min_campeonatos', 3),
                    key="min_campeonatos"
                )
            
            with col2:
                analyzer.busca_config['usar_excl_apostas_a_favor'] = st.checkbox(
                    "Excluir apostas a favor de jogadores", 
                    value=analyzer.busca_config.get('usar_excl_apostas_a_favor', True),
                    help="Exclui apostas a favor de jogadores específicos com ROI negativo"
                )
                analyzer.min_entradas_config['min_apostas_a_favor'] = st.number_input(
                    "Min. entradas:", min_value=1, max_value=50, 
                    value=analyzer.min_entradas_config.get('min_apostas_a_favor', 3),
                    key="min_apostas_a_favor"
                )
            
            # Segunda linha de exclusões
            col1, col2 = st.columns(2)
            with col1:
                analyzer.busca_config['usar_excl_apostas_contra'] = st.checkbox(
                    "Excluir apostas contra jogadores", 
                    value=analyzer.busca_config.get('usar_excl_apostas_contra', True),
                    help="Exclui apostas contra jogadores específicos com ROI negativo"
                )
                analyzer.min_entradas_config['min_apostas_contra'] = st.number_input(
                    "Min. entradas:", min_value=1, max_value=50, 
                    value=analyzer.min_entradas_config.get('min_apostas_contra', 3),
                    key="min_apostas_contra"
                )
            
            with col2:
                analyzer.busca_config['usar_excl_confrontos'] = st.checkbox(
                    "Excluir confrontos prejudiciais", 
                    value=analyzer.busca_config.get('usar_excl_confrontos', True),
                    help="Exclui confrontos específicos com ROI negativo"
                )
                analyzer.min_entradas_config['min_confrontos'] = st.number_input(
                    "Min. entradas:", min_value=1, max_value=50, 
                    value=analyzer.min_entradas_config.get('min_confrontos', 3),
                    key="min_confrontos"
                )
            
            # Terceira linha de exclusões - Times
            col1, col2 = st.columns(2)
            with col1:
                analyzer.busca_config['usar_excl_times_a_favor'] = st.checkbox(
                    "Excluir apostas a favor de times", 
                    value=analyzer.busca_config.get('usar_excl_times_a_favor', True),
                    help="Exclui apostas a favor de times específicos com ROI negativo"
                )
                analyzer.min_entradas_config['min_times_a_favor'] = st.number_input(
                    "Min. entradas:", min_value=1, max_value=50, 
                    value=analyzer.min_entradas_config.get('min_times_a_favor', 3),
                    key="min_times_a_favor"
                )
            
            with col2:
                analyzer.busca_config['usar_excl_times_contra'] = st.checkbox(
                    "Excluir apostas contra times", 
                    value=analyzer.busca_config.get('usar_excl_times_contra', True),
                    help="Exclui apostas contra times específicos com ROI negativo"
                )
                analyzer.min_entradas_config['min_times_contra'] = st.number_input(
                    "Min. entradas:", min_value=1, max_value=50, 
                    value=analyzer.min_entradas_config.get('min_times_contra', 3),
                    key="min_times_contra"
                )
            
            st.markdown("---")
            
            # Filtros por Tipo
            st.markdown("### 🎲 Filtros por Tipo")
            
            col1, col2 = st.columns(2)
            with col1:
                analyzer.busca_config['usar_excl_tipo_apostas'] = st.checkbox(
                    "Filtrar por tipo (Favorito/Azarão)", 
                    value=analyzer.busca_config.get('usar_excl_tipo_apostas', True),
                    help="Permite excluir apostas favoritas ou de azarão"
                )
                analyzer.min_entradas_config['min_tipo_apostas'] = st.number_input(
                    "Min. entradas:", min_value=1, max_value=50, 
                    value=analyzer.min_entradas_config.get('min_tipo_apostas', 3),
                    key="min_tipo_apostas"
                )
            
            with col2:
                analyzer.busca_config['usar_excl_tipo_local'] = st.checkbox(
                    "Filtrar por local (Mandante/Visitante)", 
                    value=analyzer.busca_config.get('usar_excl_tipo_local', True),
                    help="Permite excluir apostas de mandante ou visitante"
                )
                analyzer.min_entradas_config['min_tipo_local'] = st.number_input(
                    "Min. entradas:", min_value=1, max_value=50, 
                    value=analyzer.min_entradas_config.get('min_tipo_local', 3),
                    key="min_tipo_local"
                )
            
            st.markdown("---")
            
            # Filtros de Placar
            st.markdown("### ⚽ Filtros de Placar")
            
            col1, col2 = st.columns(2)
            with col1:
                analyzer.busca_config['usar_diferenca_placar_min'] = st.checkbox(
                    "Ajustar diferença de placar mínima", 
                    value=analyzer.busca_config.get('usar_diferenca_placar_min', True),
                    help="Define diferença mínima de gols para incluir a aposta"
                )
                if 'min_diferenca_placar' not in analyzer.min_entradas_config:
                    analyzer.min_entradas_config['min_diferenca_placar'] = 10
                analyzer.min_entradas_config['min_diferenca_placar'] = st.number_input(
                    "Min. entradas:", min_value=1, max_value=50, 
                    value=analyzer.min_entradas_config.get('min_diferenca_placar', 10),
                    key="min_diferenca_placar"
                )
            
            with col2:
                analyzer.busca_config['usar_diferenca_placar_max'] = st.checkbox(
                    "Ajustar diferença de placar máxima", 
                    value=analyzer.busca_config.get('usar_diferenca_placar_max', True),
                    help="Define diferença máxima de gols para incluir a aposta"
                )
            
            st.markdown("---")
            
            # Botões de ação
            col1, col2, col3 = st.columns(3)
            
            with col1:
                if st.button("✅ Selecionar Todos", use_container_width=True):
                    for key in analyzer.busca_config:
                        analyzer.busca_config[key] = True
                    st.rerun()
            
            with col2:
                if st.button("❌ Desmarcar Todos", use_container_width=True):
                    for key in analyzer.busca_config:
                        analyzer.busca_config[key] = False
                    st.rerun()
            
            with col3:
                if st.button("💾 Salvar Config", use_container_width=True):
                    if analyzer.salvar_configuracoes():
                        st.success("✅ Salvo!")
                    else:
                        st.error("❌ Erro!")
        
        with tab1:
            # Conteúdo principal na aba básica
            if not st.session_state.get('file_uploaded', False):
                st.markdown("""
                <div style="text-align: center; padding: 3rem; background: linear-gradient(45deg, #f8f9ff, #e8eeff); border-radius: 15px; margin: 2rem 0;">
                    <h3 style="color: #667eea;">🚀 Bem-vindo ao Handicap/ML Pro!</h3>
                    <p style="color: #666; font-size: 1.1rem;">Faça upload de um arquivo Excel para começar a análise avançada</p>
                    <p style="color: #888;">✨ Utilize busca otimizada para maximizar seu ROI no TipManager</p>
                </div>
                """, unsafe_allow_html=True)
                return
            
            # Métricas principais
            if analyzer.df is not None:
                base_roi = analyzer.calcular_roi(analyzer.df)
                
                col1, col2 = st.columns(2)
                
                with col1:
                    st.markdown(f"""
                    <div class="metric-card">
                        <div class="metric-value">{len(analyzer.df):,}</div>
                        <div class="metric-label">Total de Entradas</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                with col2:
                    roi_color = "#059669" if base_roi > 0 else "#dc2626"
                    st.markdown(f"""
                    <div class="metric-card">
                        <div class="metric-value" style="color: {roi_color};">{base_roi*100:.2f}%</div>
                        <div class="metric-label">ROI Base</div>
                    </div>
                    """, unsafe_allow_html=True)
            
            # Prévia dos dados
            if analyzer.df is not None:
                with st.expander("👁️ Prévia dos Dados", expanded=False):
                    st.dataframe(analyzer.df.head(100), use_container_width=True)
            
            # Configurações de análise
            st.markdown('<div class="section-header">🎯 Configurações da Análise</div>', unsafe_allow_html=True)
            
            if analyzer.df is not None:
                # Formulário de configuração
                opcoes = analyzer.obter_opcoes_formulario()
                
                col1, col2 = st.columns(2)
                
                with col1:
                    torneio_escolhido = st.selectbox(
                        "🏟️ Torneio:",
                        options=opcoes.get('torneios', ['Todos os torneios']),
                        index=len(opcoes.get('torneios', ['Todos os torneios']))-1 if opcoes.get('torneios') else 0
                    )
                    
                    campeonatos = analyzer.obter_campeonatos(torneio_escolhido)
                    campeonato_escolhido = st.selectbox(
                        "🏆 Campeonato:",
                        options=campeonatos,
                        index=len(campeonatos)-1 if campeonatos else 0
                    )
                
                with col2:
                    tips_disponiveis = analyzer.obter_tips_disponiveis(torneio_escolhido, campeonato_escolhido)
                    tip_escolhido = None
                    if tips_disponiveis:
                        tip_escolhido = st.selectbox(
                            "⚽ Tipo de Tip:",
                            options=tips_disponiveis,
                            index=len(tips_disponiveis)-1 if tips_disponiveis else 0
                        )
                    
                    roi_desejado_pct = st.number_input(
                        "📈 ROI Desejado (%):",
                        min_value=0.0,
                        max_value=100.0,
                        value=15.0,
                        step=0.5,
                        help="Meta de ROI para a otimização"
                    )
                
                # Botão de análise
                st.markdown("<br>", unsafe_allow_html=True)
                
                col_btn1, col_btn2, col_btn3 = st.columns([1, 2, 1])
                with col_btn2:
                    if st.button("🚀 Iniciar Análise", use_container_width=True, type="primary"):
                        with st.spinner("🔄 Executando busca ..."):
                            success, message = analyzer.iniciar_analise(torneio_escolhido, campeonato_escolhido, tip_escolhido, roi_desejado_pct)
                        
                        if success:
                            st.success(f"✅ {message}")
                            st.session_state['analise_completa'] = True
                            st.rerun()
                        else:
                            st.error(f"❌ {message}")
    
    # Resultados ficam fora das abas
    
    # Resultados da análise
    if st.session_state.get('analise_completa', False) and analyzer.etapas_filtros:
        st.markdown('<div class="section-header">📊 Resultados da Otimização</div>', unsafe_allow_html=True)
        
        # Tabela de etapas
        etapas_data = []
        for etapa in analyzer.etapas_filtros:
            etapas_data.append({
                'Etapa': etapa['numero'],
                'Ajuste Aplicado': etapa['ajuste'][:50] + "..." if len(etapa['ajuste']) > 50 else etapa['ajuste'],
                'Apostas': f"{etapa['entradas']:,}",
                'Lucro': f"{etapa['lucro']:.2f}",
                'ROI (%)': f"{etapa['roi']*100:.2f}%"
            })
        
        df_etapas = pd.DataFrame(etapas_data)
        
        # Aplicar cores baseado no ROI
        def highlight_roi(row):
            roi_val = float(row['ROI (%)'].rstrip('%'))
            if roi_val > 0:
                return [''] * (len(row) - 1) + ['background-color: #d1fae5; color: #065f46']
            else:
                return [''] * (len(row) - 1) + ['background-color: #fecaca; color: #991b1b']
        
        styled_df = df_etapas.style.apply(highlight_roi, axis=1)
        st.dataframe(styled_df, use_container_width=True, hide_index=True)
        
        # Seleção de etapa para relatório
        st.markdown('<div class="section-header">📄 Gerar Relatório</div>', unsafe_allow_html=True)
        
        col1, col2 = st.columns([2, 1])
        
        with col1:
            etapa_options = [f"Etapa {e['numero']} - ROI: {e['roi']*100:.2f}%" for e in analyzer.etapas_filtros]
            etapa_selecionada_idx = st.selectbox(
                "Escolha a etapa para gerar relatório:",
                range(len(etapa_options)),
                format_func=lambda x: etapa_options[x],
                index=len(etapa_options)-1
            )
        
        with col2:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("📊 Gerar Relatório", use_container_width=True, type="secondary"):
                with st.spinner("📝 Gerando relatório..."):
                    excel_data, txt_data = analyzer.gerar_relatorio_excel(etapa_selecionada_idx)
                    
                    if excel_data and txt_data:
                        st.session_state['excel_data'] = excel_data
                        st.session_state['txt_data'] = txt_data
                        st.session_state['etapa_relatorio'] = etapa_selecionada_idx
                        st.success("✅ Relatório gerado!")
        
        # Downloads
        if 'excel_data' in st.session_state and 'txt_data' in st.session_state:
            st.markdown("### 📥 Downloads")
            
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.download_button(
                    label="📊 Baixar Excel",
                    data=st.session_state['excel_data'],
                    file_name=f"Analise_Handicap_Etapa_{st.session_state['etapa_relatorio']}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            
            with col2:
                st.download_button(
                    label="📄 Baixar TXT",
                    data=st.session_state['txt_data'],
                    file_name=f"Config_Etapa_{st.session_state['etapa_relatorio']}.txt",
                    mime="text/plain",
                    use_container_width=True
                )
            
            with col3:
                if st.button("🔄 Reiniciar", help="Reiniciar estado da aplicação", use_container_width=True):
                    analyzer.reset_state()
                    if 'file_uploaded' in st.session_state:
                        del st.session_state['file_uploaded']
                    if 'excel_data' in st.session_state:
                        del st.session_state['excel_data']
                    if 'txt_data' in st.session_state:
                        del st.session_state['txt_data']
                    if 'analise_completa' in st.session_state:
                        del st.session_state['analise_completa']
                    st.rerun()

if __name__ == "__main__":
    main()
